"""
export_to_tracker.py
────────────────────
Syncs your job_pipeline DB into Job_Application_Tracker.xlsx.

Usage (from your AI folder):
    python -m job_pipeline.export_to_tracker
    python -m job_pipeline.export_to_tracker --all        # include pending_review too
    python -m job_pipeline.export_to_tracker --output "My Tracker.xlsx"

The script is SAFE TO RUN REPEATEDLY — it matches rows by apply_url
and updates in-place rather than appending duplicates.

Requirements:  pip install openpyxl psycopg2-binary
"""

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

# ── Resolve paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent          # job_pipeline/
AI_DIR       = SCRIPT_DIR.parent                        # AI/
DEFAULT_OUT  = AI_DIR / "Job_Application_Tracker.xlsx"

# ── Status mapping: pipeline → tracker ────────────────────────────────────────
# Your pipeline uses granular machine-states; the tracker uses human-readable
# stages that match how hiring managers think.
STATUS_MAP: Dict[str, str] = {
    "pending_review":  "Applied",        # scored, in your review queue
    "drafted":         "Applied",        # cover letter drafting in progress
    "approved":        "Applied",        # approved, package not yet built
    "package_ready":   "Applied",        # ready to submit
    "submitted":       "Applied",        # application sent
    "responded":       "Interview",      # employer reached out
    "rejected":        "Rejected",
    "closed":          "Withdrawn",      # skipped / auto-filtered
    # ingested / ranked are pre-queue — excluded by default
}

# Quality bucket → include by default?
INCLUDE_QUALITY = {"strong", "ok"}  # exclude "junk" and "weak" from tracker


def _safe_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return datetime.fromisoformat(str(val)).date()
    except Exception:
        return None


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def fetch_jobs(include_pending: bool = False) -> List[Dict[str, Any]]:
    """Query the DB and return rows ready to write into the tracker."""
    # Import your existing DB module so we reuse the same connection settings
    try:
        from job_pipeline.db import pg_connect
    except ImportError:
        sys.exit(
            "ERROR: Could not import job_pipeline.db — "
            "run this script from your AI/ folder as: python -m job_pipeline.export_to_tracker"
        )

    target_statuses = list(STATUS_MAP.keys())
    if not include_pending:
        target_statuses = [s for s in target_statuses if s != "pending_review"]

    conn = pg_connect()
    try:
        import psycopg2.extras
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    i.id                    AS item_id,
                    i.status,
                    i.fit_score,
                    i.quality_bucket,
                    i.summary_json,
                    i.outcome,
                    i.outcome_notes,
                    i.applied_at,
                    i.outcome_recorded_at,
                    i.updated_at,
                    p.company_name,
                    p.title,
                    p.location,
                    p.salary_text,
                    p.apply_url,
                    p.source
                FROM job_pipeline_items i
                JOIN job_postings p ON p.id = i.posting_id
                WHERE i.status = ANY(%s)
                ORDER BY
                    COALESCE(i.list_rank, i.fit_score, 0) DESC,
                    i.updated_at DESC
                """,
                (target_statuses,),
            )
            rows = cur.fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _extract_contact(summary_json: Any) -> str:
    """Pull recruiter/contact info from summary_json if Gemini stored it."""
    if not summary_json:
        return ""
    try:
        d = summary_json if isinstance(summary_json, dict) else json.loads(summary_json)
        for key in ("recruiter", "contact", "hiring_manager", "point_of_contact"):
            if d.get(key):
                return _safe_str(d[key])
    except Exception:
        pass
    return ""


def _build_notes(row: Dict[str, Any]) -> str:
    parts = []

    # Fit score
    fit = row.get("fit_score")
    if fit is not None:
        parts.append(f"Fit score: {fit:.0%}")

    # Quality bucket
    bucket = row.get("quality_bucket") or ""
    if bucket and bucket not in ("ok", ""):
        parts.append(f"Quality: {bucket}")

    # Source
    src = row.get("source") or ""
    if src:
        parts.append(f"Source: {src}")

    # Outcome notes from DB
    outcome_notes = _safe_str(row.get("outcome_notes"))
    if outcome_notes:
        parts.append(outcome_notes)

    # Gemini verdict / why snippet from summary_json
    try:
        sj = row.get("summary_json")
        d = sj if isinstance(sj, dict) else json.loads(sj or "{}")
        why = d.get("why") or d.get("reason") or d.get("verdict_reason") or ""
        if why:
            parts.append(_safe_str(why)[:120])
    except Exception:
        pass

    return " | ".join(p for p in parts if p)


def _follow_up_date(row: Dict[str, Any]) -> Optional[date]:
    """Suggest a follow-up date: 7 days after applied_at for submitted jobs."""
    status = row.get("status", "")
    applied = _safe_date(row.get("applied_at"))
    if status == "submitted" and applied:
        from datetime import timedelta
        return applied + timedelta(days=7)
    return None


def _response_date(row: Dict[str, Any]) -> Optional[date]:
    if row.get("outcome_recorded_at"):
        return _safe_date(row["outcome_recorded_at"])
    if row.get("status") in ("responded", "rejected"):
        return _safe_date(row.get("updated_at"))
    return None


def _tracker_status(row: Dict[str, Any]) -> str:
    # If outcome is recorded, prefer that over generic status map
    outcome = (row.get("outcome") or "").lower()
    if outcome in ("offer",):
        return "Offer"
    if outcome in ("interview",):
        return "Interview"
    if outcome in ("final_round", "final round"):
        return "Final Round"
    if outcome in ("rejected", "rejection"):
        return "Rejected"
    return STATUS_MAP.get(row.get("status", ""), "Applied")


# ── Excel I/O ──────────────────────────────────────────────────────────────────
# Columns in the Applications sheet (1-indexed, col A = padding):
# B=Company, C=Job Title, D=Dept/Team, E=Status, F=Date Applied,
# G=Salary Range, H=Follow-Up Date, I=Response Date, J=Contact/Recruiter,
# K=Job URL, L=Notes

COL = {
    "company":     2,
    "title":       3,
    "department":  4,
    "status":      5,
    "date_applied":6,
    "salary":      7,
    "followup":    8,
    "response":    9,
    "contact":     10,
    "url":         11,
    "notes":       12,
}
HEADER_ROW = 1
DATA_START  = 2


def load_existing_urls(ws) -> Dict[str, int]:
    """Return {apply_url: row_number} for rows already in the sheet."""
    existing: Dict[str, int] = {}
    for row in ws.iter_rows(min_row=DATA_START):
        url_cell = row[COL["url"] - 1]
        company_cell = row[COL["company"] - 1]
        if url_cell.value and str(url_cell.value).startswith("http"):
            existing[str(url_cell.value).strip()] = url_cell.row
        elif not company_cell.value:
            # empty row — stop scanning
            break
    return existing


def find_first_empty_row(ws) -> int:
    for row in ws.iter_rows(min_row=DATA_START):
        if not row[COL["company"] - 1].value:
            return row[0].row
    return ws.max_row + 1


def write_row(ws, row_num: int, data: Dict[str, Any]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    GRAY_LIGHT = "F2F2F2"
    WHITE      = "FFFFFF"
    fill_color = GRAY_LIGHT if row_num % 2 == 0 else WHITE

    def s(col_key):
        return ws.cell(row=row_num, column=COL[col_key])

    def fmt(c, val, is_date=False):
        c.value = val
        c.font  = Font(name="Arial", size=10)
        c.fill  = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        side = Side(style="thin", color="BFBFBF")
        c.border = Border(left=side, right=side, top=side, bottom=side)
        if is_date and val:
            c.number_format = "MMM D, YYYY"

    fmt(s("company"),      data["company"])
    fmt(s("title"),        data["title"])
    fmt(s("department"),   data.get("location") or "")
    fmt(s("status"),       data["tracker_status"])
    fmt(s("date_applied"), data.get("applied_date"),  is_date=True)
    fmt(s("salary"),       data.get("salary") or "")
    fmt(s("followup"),     data.get("followup_date"), is_date=True)
    fmt(s("response"),     data.get("response_date"), is_date=True)
    fmt(s("contact"),      data.get("contact") or "")
    fmt(s("url"),          data.get("url") or "")
    fmt(s("notes"),        data.get("notes") or "")


def sync_to_excel(jobs: List[Dict[str, Any]], output_path: Path) -> None:
    from openpyxl import load_workbook

    if not output_path.exists():
        sys.exit(f"ERROR: Tracker not found at {output_path}\nRun the tracker builder first.")

    wb = load_workbook(str(output_path))

    if "Applications" not in wb.sheetnames:
        sys.exit("ERROR: 'Applications' sheet not found in the tracker file.")

    ws = wb["Applications"]
    existing_urls = load_existing_urls(ws)

    added   = 0
    updated = 0

    for job in jobs:
        # Skip junk quality unless it's already submitted/responded
        bucket = (job.get("quality_bucket") or "ok").lower()
        status = job.get("status", "")
        if bucket not in INCLUDE_QUALITY and status not in ("submitted", "responded", "rejected"):
            continue

        tracker_status = _tracker_status(job)
        url = _safe_str(job.get("apply_url"))

        data = {
            "company":        _safe_str(job.get("company_name")),
            "title":          _safe_str(job.get("title")),
            "location":       _safe_str(job.get("location")),
            "tracker_status": tracker_status,
            "applied_date":   _safe_date(job.get("applied_at")),
            "salary":         _safe_str(job.get("salary_text")),
            "followup_date":  _follow_up_date(job),
            "response_date":  _response_date(job),
            "contact":        _extract_contact(job.get("summary_json")),
            "url":            url,
            "notes":          _build_notes(job),
        }

        if url and url in existing_urls:
            # Update status + notes in existing row
            row_num = existing_urls[url]
            ws.cell(row=row_num, column=COL["status"]).value = tracker_status
            ws.cell(row=row_num, column=COL["notes"]).value  = data["notes"]
            if data["response_date"]:
                ws.cell(row=row_num, column=COL["response"]).value = data["response_date"]
                ws.cell(row=row_num, column=COL["response"]).number_format = "MMM D, YYYY"
            updated += 1
        else:
            row_num = find_first_empty_row(ws)
            write_row(ws, row_num, data)
            if url:
                existing_urls[url] = row_num
            added += 1

    wb.save(str(output_path))
    print(f"\n✅ Tracker updated: {output_path.name}")
    print(f"   {added} new rows added  |  {updated} existing rows updated")
    print(f"   Total jobs processed: {len(jobs)}")


# ── CLI ────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Sync your job pipeline DB into Job_Application_Tracker.xlsx"
    )
    parser.add_argument(
        "--all", action="store_true",
        help="Also include 'pending_review' jobs (not yet submitted)"
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUT),
        help=f"Path to tracker file (default: {DEFAULT_OUT})"
    )
    args = parser.parse_args()

    output_path = Path(args.output)
    include_pending = args.all

    print(f"🔍 Fetching jobs from pipeline DB{'  (including pending_review)' if include_pending else ''}...")
    jobs = fetch_jobs(include_pending=include_pending)
    print(f"   Found {len(jobs)} jobs to process")

    if not jobs:
        print("   Nothing to sync — your pipeline may be empty or DB may be unreachable.")
        return

    sync_to_excel(jobs, output_path)


if __name__ == "__main__":
    main()
